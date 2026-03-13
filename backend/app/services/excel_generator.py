"""Excel report generator using openpyxl."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from app.services.report_service import ReportData


def _fmt_minutes(minutes: int) -> str:
    h, m = divmod(minutes, 60)
    if h and m:
        return f"{h}h {m}m"
    elif h:
        return f"{h}h"
    return f"{m}m"


class ExcelGenerator:

    PRIMARY_HEX = "2196F3"
    SUCCESS_HEX = "4CAF50"
    LIGHT_HEX = "F5F5F5"
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(name="Calibri", size=10)
    HEADER_FILL = PatternFill("solid", fgColor="2196F3")
    ALT_FILL = PatternFill("solid", fgColor="F5F5F5")
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    def generate(self, data: ReportData, output_path: str) -> str:
        wb = Workbook()

        self._build_summary_sheet(wb, data)
        self._build_projects_sheet(wb, data)
        self._build_tasks_sheet(wb, data)
        self._build_time_detail_sheet(wb, data)
        if data.persons:
            self._build_persons_sheet(wb, data)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        return output_path

    def _apply_header(self, ws, row, headers, col_widths):
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = w

    def _build_summary_sheet(self, wb, data: ReportData):
        ws = wb.create_sheet("Riepilogo")
        ws.sheet_view.showGridLines = False

        ws["A1"] = data.title
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=self.PRIMARY_HEX)
        ws["A2"] = f"Periodo: {data.period_from} - {data.period_to}"
        ws["A2"].font = Font(name="Calibri", size=10, color="757575")
        ws["A3"] = f"Generato il: {data.generated_at[:16].replace('T', ' ')}"
        ws["A3"].font = Font(name="Calibri", size=10, color="757575")

        kpis = [
            ("Ore lavorate", _fmt_minutes(data.total_logged_minutes)),
            ("Task completati", data.total_done_tasks),
            ("Task aperti", data.total_open_tasks),
            ("% Completamento", f"{data.avg_completion_pct:.0f}%"),
        ]
        for col, (label, value) in enumerate(kpis, start=1):
            ws.cell(row=5, column=col, value=label).font = Font(
                name="Calibri", bold=True, size=9, color="757575"
            )
            ws.cell(row=6, column=col, value=value).font = Font(
                name="Calibri", bold=True, size=18, color=self.PRIMARY_HEX
            )
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Trend section
        ws["A8"] = "Confronto periodo precedente"
        ws["A8"].font = Font(name="Calibri", bold=True, size=12)

        headers = ["Metrica", "Corrente", "Precedente", "Variazione %"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        rows = [
            ["Ore lavorate", _fmt_minutes(data.total_logged_minutes),
             _fmt_minutes(data.prev_logged_minutes), f"{data.trend_minutes_pct:+.1f}%"],
            ["Task completati", data.total_done_tasks,
             data.prev_done_tasks, f"{data.trend_tasks_pct:+.1f}%"],
        ]
        for r, row in enumerate(rows, start=10):
            fill = self.ALT_FILL if r % 2 == 0 else PatternFill()
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = self.NORMAL_FONT
                cell.fill = fill
                cell.border = self.THIN_BORDER

    def _build_projects_sheet(self, wb, data: ReportData):
        ws = wb.create_sheet("Progetti")
        ws.sheet_view.showGridLines = False

        headers = ["Progetto", "Cliente", "Stato", "Task totali",
                   "Task completati", "Task aperti", "Ore lavorate",
                   "Ore stimate", "% Completamento"]
        col_widths = [30, 20, 12, 12, 15, 12, 14, 14, 16]
        self._apply_header(ws, 1, headers, col_widths)

        # Hidden column for chart (raw minutes)
        ws.column_dimensions[get_column_letter(10)].hidden = True
        ws.cell(row=1, column=10, value="Minuti (raw)")

        for r, p in enumerate(data.projects, start=2):
            fill = self.ALT_FILL if r % 2 == 0 else PatternFill()
            values = [
                p.name, p.client_name or "", p.status,
                p.total_tasks, p.done_tasks, p.open_tasks,
                _fmt_minutes(p.logged_minutes),
                _fmt_minutes(p.estimated_minutes) if p.estimated_minutes else "",
                p.completion_pct / 100,
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = self.NORMAL_FONT
                cell.fill = fill
                cell.border = self.THIN_BORDER
                if c == 9:
                    cell.number_format = "0%"
            # Raw minutes for chart
            ws.cell(row=r, column=10, value=p.logged_minutes)

        # Bar chart using raw minutes
        n = len(data.projects)
        if n > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Ore per progetto"
            chart.y_axis.title = "Minuti"
            chart.grouping = "clustered"
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=10, max_col=10, min_row=1, max_row=n + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, f"A{n + 4}")

    def _build_tasks_sheet(self, wb, data: ReportData):
        ws = wb.create_sheet("Task")
        ws.sheet_view.showGridLines = False

        headers = ["Progetto", "Task", "Stato", "Completato il", "Ore log", "Ore stimate"]
        col_widths = [25, 40, 12, 15, 12, 14]
        self._apply_header(ws, 1, headers, col_widths)

        r = 2
        for p in data.projects:
            for t in p.tasks:
                fill = self.ALT_FILL if r % 2 == 0 else PatternFill()
                values = [
                    p.name, t.title, t.status,
                    t.completed_at or "",
                    _fmt_minutes(t.logged_minutes),
                    _fmt_minutes(t.estimated_minutes) if t.estimated_minutes else "",
                ]
                for c, val in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = self.NORMAL_FONT
                    cell.fill = fill
                    cell.border = self.THIN_BORDER
                r += 1

    def _build_time_detail_sheet(self, wb, data: ReportData):
        ws = wb.create_sheet("Log ore")
        ws.sheet_view.showGridLines = False

        headers = ["Data", "Progetto", "Task", "Ore", "Nota"]
        col_widths = [12, 25, 35, 10, 40]
        self._apply_header(ws, 1, headers, col_widths)

        r = 2
        for log in data.time_log_details:
            fill = self.ALT_FILL if r % 2 == 0 else PatternFill()
            values = [log.logged_at, log.project_name, log.task_title,
                      _fmt_minutes(log.minutes), log.note or ""]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = self.NORMAL_FONT
                cell.fill = fill
                cell.border = self.THIN_BORDER
            r += 1

    def _build_persons_sheet(self, wb, data: ReportData):
        ws = wb.create_sheet("Per persona")
        ws.sheet_view.showGridLines = False

        headers = ["Persona", "Ore lavorate", "Task completati", "Progetti"]
        col_widths = [25, 15, 16, 40]
        self._apply_header(ws, 1, headers, col_widths)

        for r, person in enumerate(data.persons, start=2):
            fill = self.ALT_FILL if r % 2 == 0 else PatternFill()
            values = [
                person.display_name,
                _fmt_minutes(person.logged_minutes),
                person.done_tasks,
                ", ".join(person.projects),
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = self.NORMAL_FONT
                cell.fill = fill
                cell.border = self.THIN_BORDER
